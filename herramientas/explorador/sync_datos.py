#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sincroniza el Explorador de Proyectos con la hoja maestra de Google Drive.

    "base de datos de nattiva locker"
    fileId 1s0RBSd4A8r2ck-KuLkLABuXx2ZrgpYahBEP9IQxdwng

REGLA DE NEGOCIO
    Hoja 1 ("Cotizador Nattiva_Online_Tabla") es la principal: manda siempre.
    Hoja 2 ("performance") solo aporta lo que NO exista en la hoja 1.
    La hoja 2 nunca pisa un dato de la hoja 1 (sus precios son PRECIO LISTA,
    sin descuento, y su numeracion de unidades usa otro criterio).

USO
    1. Descarga la hoja de Drive como .xlsx  (Archivo > Descargar > .xlsx)
    2. python3 sync_datos.py ruta/al/archivo.xlsx
    3. Revisa el reporte, sobre todo los AVISOS.
    4. git add -A && git commit && git push

Reescribe:  index.html (el bloque `var DB=` y los conteos por desarrolladora)
            nattiva_data_limpia.csv
"""

import sys, os, re, csv, json, unicodedata, collections

HERE = os.path.dirname(os.path.abspath(__file__))
HTML = os.path.join(HERE, 'index.html')
CSV_OUT = os.path.join(HERE, 'nattiva_data_limpia.csv')

HOJA1 = 'Cotizador Nattiva_Online_Tabla'   # se busca por prefijo, tolera espacios
HOJA2 = 'performance'

# ---------------------------------------------------------------------------
# CONFIGURACION EDITABLE
# ---------------------------------------------------------------------------

# La hoja 2 escribe algunos nombres distinto que la hoja 1. Sin estos alias,
# el mismo edificio entraria dos veces como si fuera un proyecto nuevo.
# Clave y valor van normalizados (mayusculas, sin acentos, sin espacios).
ALIAS_DEV = {
    'VYV': 'V&V',
    'GRANADEROINMOBILIARIA': 'GRANADERO',
    'NULL': 'POR CONFIRMAR',
}

# (dev_norm, proyecto_norm) de la hoja 2  ->  nombre de proyecto en la hoja 1
ALIAS_PROYECTO = {
    ('VYV', 'SANTACATALINA'): 'TORRE SANTA CATALINA',
    ('VYV', 'SANFELIPE'): 'SAN FELIPE VYV',
}

# Casos donde un proyecto de la hoja 2 son en realidad dos de la hoja 1,
# separados por la columna TORRE.  (dev_norm, proyecto_norm) -> {torre: nombre}
ALIAS_PROYECTO_POR_TORRE = {
    ('VYV', 'VILLARAN'): {'1': 'VILLARÁN 1', '2': 'VILLARÁN 2'},
}

# Nombre definitivo de un proyecto, se aplique en la hoja que se aplique.
# Sirve para celdas que Google exporta distinto segun el formato: "28 DE JULIO
# 0360" sale como "28 DE julio 0360" en CSV y "28 DE July 0360" en XLSX porque
# la celda esta autoformateada como fecha. Conviene arreglarla en la hoja
# (formato > texto sin formato); mientras tanto, esto la deja estable.
NOMBRE_CANONICO = {
    ('GRATTOINMOBILIARIA', '28DEJULY360'): '28 DE JULIO 360',
    ('GRATTOINMOBILIARIA', '28DEJULIO360'): '28 DE JULIO 360',
}

ALIAS_DISTRITO = {
    'MAGDALENA': 'Magdalena Del Mar',
    'SANTIAGODESURCO': 'Surco',
    'NULL': 'Por Confirmar',
}

# Limpiezas puntuales heredadas de la version original del explorador.
OVR_VISTA = {
    'PENDIENTE': 'Por confirmar',
    'null': 'Por confirmar',
    'NA': 'Por confirmar',
    '': 'Por confirmar',
    'vista': 'Exterior',
    'VISTA CALLE E INTERNA': 'Vista Calle',
    'CAMINOS DEL INCA (PARQUE DE LA AMISTAD)': 'Caminos Del Inca (parque De La Amistad)',
}
OVR_BANCO = {
    'BCP': 'BCP', 'BBVA': 'BBVA', 'BANBIF': 'BanBif',
    'T. BANCOS': 'Todos los bancos', 'BANBIF / BBVA': 'BANBIF / BBVA',
    'IBK - BBVA': 'IBK - BBVA',
    'PENDIENTE': 'Por definir', 'POR DEFINIR': 'Por definir',
    'null': 'Por definir', '': 'Por definir', 'NA': 'Por definir',
}
OVR_DORM = {'PENDIENTE': -1, 'NA': -1, 'null': -1, '': -1, 'MONOAMBIENTE': 0}
OVR_MONEDA = {'SOLES': 0, 'DÓLARES': 1, 'DOLARES': 1, 'null': 0, '': 0}
OVR_VIS = {'SI': 1, 'NO': 0, 'null': -1, '': -1, 'NA': -1}

# Corrige el ano de entrega cuando la fecha por unidad de la hoja esta mal.
ENTREGA_FIX = {'BENAVIDES 1361': '2028'}

# Si la hoja 2 quiere agregar mas de este % del tamano que el proyecto tiene en
# la hoja 1, se avisa: casi siempre significa que numeran distinto.
AVISO_UMBRAL = 0.25

MESES = {1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
         7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Oct', 11: 'Nov', 12: 'Dic'}

# ---------------------------------------------------------------------------


def norm(x):
    """Clave de comparacion: mayusculas, sin acentos, solo letras y digitos."""
    x = str(x if x is not None else '').strip().upper()
    if re.fullmatch(r'-?\d+\.0', x):
        x = x[:-2]
    x = unicodedata.normalize('NFD', x)
    x = ''.join(c for c in x if unicodedata.category(c) != 'Mn')
    return re.sub(r'[^A-Z0-9]', '', x.replace('&', 'Y'))


def norm_proj(x):
    """Como norm(), pero sin ceros a la izquierda en cada numero:
    28DEJULIO0360 == 28DEJULIO360, y SALAVERRYAVENUE2205 no se toca."""
    return re.sub(r'0*(\d+)', lambda m: m.group(1), norm(x))


def base_unidad(x):
    """Numero de unidad sin la torre: 'H-1103' -> '1103', '1105B' -> '1105'."""
    u = norm(x)
    d = re.findall(r'\d+', u)
    return (d[-1].lstrip('0') or '0') if d else u


def canonico(dev, proyecto):
    """Nombre definitivo del proyecto, igual venga de la hoja que venga."""
    return NOMBRE_CANONICO.get((norm(dev), norm_proj(proyecto)), proyecto)


def tc(s):
    """Capitaliza cada palabra, tambien despues de '.' y '/'."""
    out, cap = [], True
    for ch in str(s):
        out.append(ch.upper() if cap and ch.isalpha() else (ch.lower() if ch.isalpha() else ch))
        cap = not ch.isalnum()
    return ''.join(out)


def txt(v):
    s = '' if v is None else str(v).strip()
    return s[:-2] if re.fullmatch(r'-?\d+\.0', s) else s


def num(v, d=0.0):
    try:
        return float(str(v).replace(',', '').strip())
    except (TypeError, ValueError):
        return d


def compacta(x):
    return int(x) if float(x) == int(float(x)) else float(x)


def fecha_entrega(v):
    v = txt(v)
    m = re.fullmatch(r'(\d{4})-(\d{2})-(\d{2})', v)
    if m:
        return '%s-%s' % (MESES[int(m.group(2))], m.group(1)[2:])
    m = re.fullmatch(r'(\d{4})-(\d{2})-(\d{2}) .*', v)
    if m:
        return '%s-%s' % (MESES[int(m.group(2))], m.group(1)[2:])
    m = re.fullmatch(r'([A-Za-zÁÉÍÓÚáéíóú]{3})-(\d{2})', v)
    if m:
        return m.group(1).capitalize() + '-' + m.group(2)
    return v or 'null'


def vista(v):
    v = txt(v)
    return OVR_VISTA.get(v, OVR_VISTA.get(v.upper(), tc(v)))


def banco(v):
    v = txt(v)
    return OVR_BANCO.get(v, OVR_BANCO.get(v.upper(), tc(v)))


def dormitorios(v):
    v = txt(v)
    if v.upper() in OVR_DORM:
        return OVR_DORM[v.upper()]
    try:
        return compacta(float(v.replace(',', '.')))
    except ValueError:
        return -1


# ---------------------------------------------------------------------------
# Lectura
# ---------------------------------------------------------------------------

def lee_hojas(path):
    try:
        import openpyxl
    except ImportError:
        sys.exit('Falta openpyxl:  pip install openpyxl')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    def hoja(prefijo):
        for nm in wb.sheetnames:
            if nm.strip().lower().startswith(prefijo.strip().lower()):
                return wb[nm]
        sys.exit('No encuentro la hoja "%s". Hojas: %s' % (prefijo, wb.sheetnames))
    def filas(ws):
        it = ws.iter_rows(values_only=True)
        hdr = [txt(h) for h in next(it)]
        out = []
        for r in it:
            if any(c not in (None, '') for c in r):
                out.append(dict(zip(hdr, r)))
        return out
    return filas(hoja(HOJA1)), filas(hoja(HOJA2))


def canon1(r):
    """Fila de la hoja 1 -> registro canonico."""
    return {
        'distrito': tc(txt(r.get('DISTRITO'))) or 'Por Confirmar',
        'dev': txt(r.get('DESARROLLADORA')).upper() or 'POR CONFIRMAR',
        'proyecto': canonico(txt(r.get('DESARROLLADORA')).upper(), txt(r.get('PROYECTO'))),
        'unidad': txt(r.get('N° INMUEBLE')),
        'modelo': txt(r.get('MODELO')),
        'at': num(r.get('AREA TECHADA')), 'al': num(r.get('AREA LIBRE')),
        'total': num(r.get('AREA TOTAL')),
        'vista': vista(r.get('VISTA')),
        'dorm': dormitorios(r.get('DORMITORIOS')),
        'moneda': OVR_MONEDA.get(txt(r.get('MONEDA')).upper(), 0),
        'precio': num(r.get('PRECIO')),
        'entrega': fecha_entrega(r.get('F. ENTREGA')),
        'vis': OVR_VIS.get(txt(r.get('RESTRICCIÓN VIS')).upper(), -1),
        'cuota': num(r.get('CUOTA APROXIMADA')),
        'banco': banco(r.get('BANCO')),
        'origen': 'hoja1',
    }


def canon2(r):
    """Fila de la hoja 2 -> registro canonico, con los alias aplicados.

    La hoja 2 no trae fecha de entrega, restriccion VIS, cuota ni banco:
    esos campos quedan sin definir, no se inventan.
    """
    dev_raw = txt(r.get('INMOBILIARIA'))
    dev = ALIAS_DEV.get(norm(dev_raw), dev_raw.upper()) or 'POR CONFIRMAR'
    proj_raw = txt(r.get('PROYECTO'))
    k = (norm(dev), norm_proj(proj_raw))
    proyecto = proj_raw
    if k in ALIAS_PROYECTO_POR_TORRE:
        proyecto = ALIAS_PROYECTO_POR_TORRE[k].get(txt(r.get('TORRE')), proj_raw)
    elif k in ALIAS_PROYECTO:
        proyecto = ALIAS_PROYECTO[k]
    proyecto = canonico(dev, proyecto)
    dist_raw = txt(r.get('DISTRITO'))
    return {
        'distrito': ALIAS_DISTRITO.get(norm(dist_raw), tc(dist_raw)) or 'Por Confirmar',
        'dev': dev,
        'proyecto': proyecto,
        'unidad': txt(r.get('N° Unidad')),
        'modelo': txt(r.get('MODELO')),
        'at': num(r.get('A T')), 'al': num(r.get('AL')),
        'total': num(r.get('AreaTotal')),
        'vista': vista(r.get('VISTA')),
        'dorm': dormitorios(r.get('Dormitorios')),
        'moneda': OVR_MONEDA.get(txt(r.get('MONEDA')).upper(), 0),
        'precio': num(r.get('PRECIO LISTA')),
        'entrega': 'null',
        'vis': -1,
        'cuota': 0,
        'banco': 'Por definir',
        'origen': 'hoja2',
    }


# ---------------------------------------------------------------------------
# Fusion: hoja 1 manda, hoja 2 solo rellena huecos
# ---------------------------------------------------------------------------

def fusiona(f1, f2, log):
    base = [canon1(r) for r in f1]

    exacto = collections.defaultdict(set)   # (dev, proy) -> numeros tal cual
    basico = collections.defaultdict(set)   # (dev, proy) -> numeros sin torre
    tam1 = collections.Counter()
    for c in base:
        k = (norm(c['dev']), norm_proj(c['proyecto']))
        exacto[k].add(norm(c['unidad']))
        basico[k].add(base_unidad(c['unidad']))
        tam1[k] += 1

    agregadas, por_proyecto, proyectos_nuevos = [], collections.Counter(), set()
    for r in f2:
        c = canon2(r)
        k = (norm(c['dev']), norm_proj(c['proyecto']))
        if k not in exacto:
            proyectos_nuevos.add(k)
            agregadas.append(c); por_proyecto[k] += 1
            continue
        if norm(c['unidad']) in exacto[k] or base_unidad(c['unidad']) in basico[k]:
            continue                                   # ya existe: manda la hoja 1
        agregadas.append(c); por_proyecto[k] += 1

    # Avisos: nombres de la hoja 2 que se parecen a un proyecto de la hoja 1.
    for k in sorted(proyectos_nuevos):
        parecidos = [a for a in exacto
                     if a[0] == k[0] and (k[1] in a[1] or a[1] in k[1])]
        if parecidos:
            log.aviso('"%s / %s" entra como PROYECTO NUEVO pero se parece a %s. '
                      'Si es el mismo, agregalo a ALIAS_PROYECTO.'
                      % (k[0], k[1], ', '.join(p[1] for p in parecidos)))

    for k, n in por_proyecto.items():
        if k in tam1 and n > tam1[k] * AVISO_UMBRAL:
            log.aviso('"%s / %s": la hoja 2 suma %d unidades sobre las %d de la hoja 1 '
                      '(%.0f%%). Revisa que no numeren distinto.'
                      % (k[0], k[1], n, tam1[k], 100.0 * n / tam1[k]))

    return base + agregadas, agregadas, por_proyecto, proyectos_nuevos


# ---------------------------------------------------------------------------
# Construccion del DB embebido
# ---------------------------------------------------------------------------

def construye_db(regs):
    DIST, DEV, VISTA, BANCO, ENT = [], [], [], [], []
    def idx(lst, v):
        try:
            return lst.index(v)
        except ValueError:
            lst.append(v)
            return len(lst) - 1

    grupos = collections.OrderedDict()
    for c in regs:
        grupos.setdefault((norm(c['dev']), norm_proj(c['proyecto'])), []).append(c)

    P = []
    for _, cs in grupos.items():
        nombre = cs[0]['proyecto']
        dist = collections.Counter(c['distrito'] for c in cs).most_common(1)[0][0]
        u = [[c['unidad'], c['modelo'],
              compacta(c['at']), compacta(c['al']), compacta(c['total']),
              idx(VISTA, c['vista']), c['dorm'], c['moneda'], compacta(c['precio']),
              idx(ENT, c['entrega']), c['vis'], compacta(c['cuota']),
              idx(BANCO, c['banco'])] for c in cs]
        P.append({'n': nombre, 'dv': idx(DEV, cs[0]['dev']), 'di': idx(DIST, dist),
                  'b': '', 'u': u})

    P.sort(key=lambda p: (-len(p['u']), p['n']))
    return {'DIST': DIST, 'DEV': DEV, 'VISTA': VISTA, 'BANCO': BANCO, 'ENT': ENT, 'P': P}


def escribe_html(db, log):
    s = open(HTML, encoding='utf-8').read()
    original = s

    i = s.find('var DB=')
    if i < 0:
        sys.exit('No encuentro "var DB=" en index.html')
    j = i + len('var DB=')
    d = 0
    for k in range(j, len(s)):
        if s[k] == '{':
            d += 1
        elif s[k] == '}':
            d -= 1
            if d == 0:
                break
    if s[k + 1] != ';':
        sys.exit('El bloque var DB no termina en ";"')
    s = s[:j] + json.dumps(db, ensure_ascii=False, separators=(', ', ': ')) + s[k + 1:]

    # conteo de proyectos por desarrolladora en las fichas de aliados
    pc = collections.Counter(db['DEV'][p['dv']] for p in db['P'])
    m = re.search(r'var PARTNERS=(\[.*?\]);\n', s, re.S)
    if m:
        pt = json.loads(m.group(1))
        for t in pt:
            if t.get('s') == 'dev':
                n = pc.get(t['u'], 0)
                if t.get('count') != n:
                    log.info('  conteo %s: %s -> %d' % (t['n'], t.get('count'), n))
                t['count'] = n
        s = s[:m.start(1)] + json.dumps(pt, ensure_ascii=False) + s[m.end(1):]
    else:
        log.aviso('No encuentro var PARTNERS: los conteos por desarrolladora no se tocaron.')

    # cifras escritas a mano en el texto
    tot = sum(len(p['u']) for p in db['P'])
    s, n1 = re.subn(r'Data real de [\d,]+ unidades', 'Data real de {:,} unidades'.format(tot), s)
    s, n2 = re.subn(r'los \d+ proyectos del mercado', 'los %d proyectos del mercado' % len(db['P']), s)
    if not n1 or not n2:
        log.aviso('No pude actualizar alguna cifra fija del texto (%d/%d).' % (n1, n2))

    if s == original:
        log.aviso('index.html quedo igual que antes.')
    open(HTML, 'w', encoding='utf-8').write(s)


def escribe_csv(regs):
    cols = [('DISTRITO', 'distrito'), ('DESARROLLADORA', 'dev'), ('PROYECTO', 'proyecto'),
            ('N INMUEBLE', 'unidad'), ('MODELO', 'modelo'), ('AREA TECHADA', 'at'),
            ('AREA LIBRE', 'al'), ('AREA TOTAL', 'total'), ('VISTA', 'vista'),
            ('DORMITORIOS', 'dorm'), ('MONEDA', 'moneda'), ('PRECIO', 'precio'),
            ('F ENTREGA', 'entrega'), ('VIS', 'vis'), ('CUOTA', 'cuota'),
            ('BANCO', 'banco'), ('ORIGEN', 'origen')]
    with open(CSV_OUT, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow([c[0] for c in cols])
        for c in regs:
            w.writerow([c[k] for _, k in cols])


class Log(object):
    def __init__(self):
        self.avisos = []
    def info(self, m):
        print(m)
    def aviso(self, m):
        self.avisos.append(m)


def main():
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    xlsx = sys.argv[1]
    if not os.path.exists(xlsx):
        sys.exit('No existe: %s' % xlsx)

    log = Log()
    f1, f2 = lee_hojas(xlsx)
    print('Hoja 1 (principal): %d filas' % len(f1))
    print('Hoja 2 (performance): %d filas' % len(f2))

    regs, agregadas, por_proyecto, nuevos = fusiona(f1, f2, log)

    print('\nDe la hoja 2 se agregaron %d unidades que no estaban en la hoja 1:' % len(agregadas))
    for k, n in sorted(por_proyecto.items(), key=lambda x: -x[1]):
        marca = '  [PROYECTO NUEVO]' if k in nuevos else ''
        print('   %-22s %-26s %4d%s' % (k[0], k[1], n, marca))
    if not agregadas:
        print('   (ninguna)')

    db = construye_db(regs)
    print('\nResultado: %d proyectos / %d unidades / %d desarrolladoras / %d distritos'
          % (len(db['P']), sum(len(p['u']) for p in db['P']), len(db['DEV']), len(db['DIST'])))

    print('\nEscribiendo index.html...')
    escribe_html(db, log)
    escribe_csv(regs)
    print('Escrito nattiva_data_limpia.csv (%d filas)' % len(regs))

    if log.avisos:
        print('\n%d AVISO(S) - revisalos antes de publicar:' % len(log.avisos))
        for a in log.avisos:
            print('  - %s' % a)
    else:
        print('\nSin avisos.')


if __name__ == '__main__':
    main()
